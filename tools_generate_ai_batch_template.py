from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def u(text: str) -> str:
    return text


def main() -> Path:
    repo = Path(r"d:/AI_project/auto-test-platform")
    out = repo / "AI_batch_generation_test_template.xlsx"

    headers = [
        u("\u6d4b\u8bd5\u63cf\u8ff0"),
        u("\u811a\u672c\u540d\u79f0"),
        u("\u63cf\u8ff0"),
        u("\u6807\u7b7e"),
        u("\u6a21\u5757"),
        u("\u4f18\u5148\u7ea7"),
    ]

    rows = [
        [
            u("\u6253\u5f00\u672c\u7cfb\u7edf\u767b\u5f55\u9875\uff0c\u4f7f\u7528\u7ba1\u7406\u5458\u8d26\u53f7\u767b\u5f55\uff0c\u767b\u5f55\u6210\u529f\u540e\u8fdb\u5165\u9879\u76ee\u5217\u8868\u9875\uff0c\u68c0\u67e5\u9875\u9762\u6807\u9898\u548c\u5bfc\u822a\u680f\u3002"),
            u("\u767b\u5f55-\u8fdb\u5165\u9879\u76ee\u5217\u8868"),
            u("\u9a8c\u8bc1\u767b\u5f55\u5e76\u8fdb\u5165\u9996\u9875"),
            u("\u767b\u5f55,\u5bfc\u822a,\u5192\u70df"),
            u("\u767b\u5f55"),
            "P0",
        ],
        [
            u("\u767b\u5f55\u540e\u8fdb\u5165\u9879\u76ee\u5217\u8868\u9875\uff0c\u70b9\u51fb\u65b0\u5efa\u9879\u76ee\uff0c\u586b\u5199\u540d\u79f0\u3001\u7c7b\u578b\u548c\u63cf\u8ff0\uff0c\u4fdd\u5b58\u540e\u68c0\u67e5\u5217\u8868\u3002"),
            u("\u9879\u76ee-\u65b0\u5efa"),
            u("\u9a8c\u8bc1\u9879\u76ee\u521b\u5efa\u6d41\u7a0b"),
            u("\u9879\u76ee,\u65b0\u5efa,\u6838\u5fc3"),
            u("\u9879\u76ee\u7ba1\u7406"),
            "P0",
        ],
        [
            u("\u767b\u5f55\u540e\u8fdb\u5165\u9879\u76ee\u8be6\u60c5\u9875\uff0c\u8fdb\u5165\u53d8\u91cf\u7ba1\u7406\uff0c\u65b0\u589e\u53d8\u91cf\u5e76\u4fdd\u5b58\u3002"),
            u("\u53d8\u91cf-\u65b0\u589e"),
            u("\u9a8c\u8bc1\u9879\u76ee\u53d8\u91cf\u7ef4\u62a4"),
            u("\u53d8\u91cf,\u914d\u7f6e,\u9879\u76ee"),
            u("\u53d8\u91cf\u7ba1\u7406"),
            "P1",
        ],
        [
            u("\u767b\u5f55\u540e\u8fdb\u5165\u811a\u672c\u5217\u8868\u9875\uff0c\u65b0\u5efa\u811a\u672c\u6a21\u5757\uff0c\u5185\u5bb9\u5305\u542b\u6253\u5f00\u9996\u9875\u548c\u70b9\u51fb\u767b\u5f55\u6309\u94ae\u3002"),
            u("\u811a\u672c-\u65b0\u5efa"),
            u("\u9a8c\u8bc1\u811a\u672c\u521b\u5efa\u4e0e\u4fdd\u5b58"),
            u("\u811a\u672c,\u521b\u5efa,\u6a21\u5757"),
            u("\u811a\u672c\u7ba1\u7406"),
            "P0",
        ],
        [
            u("\u767b\u5f55\u540e\u8fdb\u5165\u811a\u672c\u7f16\u8f91\u9875\uff0c\u4fee\u6539\u767b\u5f55\u6b65\u9aa4\u7684\u5b9a\u4f4d\u5668\u548c\u8f93\u5165\u503c\uff0c\u4fdd\u5b58\u540e\u518d\u6b21\u8fdb\u5165\u786e\u8ba4\u4fee\u6539\u3002"),
            u("\u811a\u672c-\u7f16\u8f91"),
            u("\u9a8c\u8bc1\u811a\u672c\u7f16\u8f91"),
            u("\u811a\u672c,\u7f16\u8f91,\u6b65\u9aa4"),
            u("\u811a\u672c\u7ba1\u7406"),
            "P1",
        ],
        [
            u("\u767b\u5f55\u540e\u8fdb\u5165\u6267\u884c\u8bb0\u5f55\u9875\uff0c\u70b9\u51fb\u6700\u8fd1\u4e00\u6b21\u6267\u884c\u7684\u8be6\u60c5\uff0c\u68c0\u67e5\u6b65\u9aa4\u72b6\u6001\u548c\u9519\u8bef\u4fe1\u606f\u3002"),
            u("\u6267\u884c-\u8be6\u60c5"),
            u("\u9a8c\u8bc1\u6267\u884c\u8bb0\u5f55"),
            u("\u6267\u884c,\u8bb0\u5f55,\u62a5\u544a"),
            u("\u6267\u884c\u7ba1\u7406"),
            "P1",
        ],
        [
            u("\u767b\u5f55\u540e\u6253\u5f00\u8ba1\u5212\u7ba1\u7406\u9875\uff0c\u65b0\u5efa\u8ba1\u5212\uff0c\u9009\u62e9\u4e24\u4e2a\u811a\u672c\u540e\u4fdd\u5b58\u3002"),
            u("\u8ba1\u5212-\u65b0\u5efa"),
            u("\u9a8c\u8bc1\u8ba1\u5212\u521b\u5efa"),
            u("\u8ba1\u5212,\u521b\u5efa,\u8c03\u5ea6"),
            u("\u8ba1\u5212\u7ba1\u7406"),
            "P0",
        ],
        [
            u("\u767b\u5f55\u540e\u8fdb\u5165\u8ba1\u5212\u8be6\u60c5\u9875\uff0c\u5f00\u542f\u5b9a\u65f6\u6267\u884c\uff0c\u8bbe\u7f6e cron \u4e3a\u6bcf\u5206\u949f\u4e00\u6b21\u3002"),
            u("\u8ba1\u5212-\u5b9a\u65f6"),
            u("\u9a8c\u8bc1\u8ba1\u5212\u5b9a\u65f6"),
            u("\u8ba1\u5212,\u5b9a\u65f6,cron"),
            u("\u8c03\u5ea6\u4e2d\u5fc3"),
            "P0",
        ],
        [
            u("\u767b\u5f55\u540e\u6253\u5f00 AI \u667a\u80fd\u5206\u6790\u5f39\u7a97\uff0c\u5bf9\u6700\u8fd1\u4e00\u6b21\u5931\u8d25\u6267\u884c\u8fdb\u884c\u5206\u6790\u3002"),
            u("AI-\u81ea\u6108"),
            u("\u9a8c\u8bc1 AI \u81ea\u6108\u5206\u6790"),
            u("AI,\u81ea\u6108,\u5206\u6790"),
            u("AI\u81ea\u6108"),
            "P0",
        ],
        [
            u("\u767b\u5f55\u540e\u6253\u5f00 AI \u6279\u91cf\u751f\u6210\u5f39\u7a97\uff0c\u5148\u7c98\u8d34\u4e09\u6761\u8bf4\u660e\uff0c\u7136\u540e\u751f\u6210\u5e76\u6838\u5bf9\u7ed3\u679c\u3002"),
            u("AI-\u6279\u91cf"),
            u("\u9a8c\u8bc1 AI \u6279\u91cf\u751f\u6210"),
            u("AI,\u6279\u91cf,\u751f\u6210"),
            u("AI\u751f\u6210"),
            "P0",
        ],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"A": 60, "B": 22, "C": 28, "D": 24, "E": 18, "F": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    guide = wb.create_sheet("Guide")
    guide.append([u("\u5b57\u6bb5"), u("\u7528\u9014"), u("\u8bf4\u660e")])
    guide_rows = [
        [u("\u6d4b\u8bd5\u63cf\u8ff0"), u("\u5fc5\u586b"), u("\u8bf7\u5199\u660e\u767b\u5f55\u7cfb\u7edf -> \u8fdb\u5165\u9875\u9762 -> \u70b9\u51fb\u64cd\u4f5c -> \u9a8c\u8bc1\u7ed3\u679c\u3002")],
        [u("\u811a\u672c\u540d\u79f0"), u("\u53ef\u9009"), u("\u4f5c\u4e3a\u751f\u6210\u811a\u672c\u7684\u9ed8\u8ba4\u540d\u79f0\u3002")],
        [u("\u63cf\u8ff0"), u("\u53ef\u9009"), u("\u4f5c\u4e3a\u811a\u672c\u8bf4\u660e\u3002")],
        [u("\u6807\u7b7e"), u("\u53ef\u9009"), u("\u5efa\u8bae\u7528\u9017\u53f7\u5206\u9694\u3002")],
        [u("\u6a21\u5757"), u("\u53ef\u9009"), u("\u7528\u6765\u533a\u5206\u767b\u5f55\u3001\u9879\u76ee\u3001\u811a\u672c\u3001\u6267\u884c\u3001\u62a5\u544a\u3001AI \u573a\u666f\u3002")],
        [u("\u4f18\u5148\u7ea7"), u("\u53ef\u9009"), u("P0 / P1 / P2 \u90fd\u53ef\u4ee5\u3002")],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    guide.freeze_panes = "A2"
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 14
    guide.column_dimensions["C"].width = 78
    for row in guide.iter_rows(min_row=2, max_row=guide.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out)
    return out


if __name__ == "__main__":
    path = main()
    print(path)
